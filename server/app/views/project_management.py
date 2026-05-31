# server/app/views/project_management.py
# -*- coding: utf-8 -*-

from flask import request, jsonify, g, send_file, make_response
from app.repositories.project_management_repo import project_management_repo
from app.decorators import api_permission_required
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def init_project_management_routes(bp):
    """初始化项目管理相关路由"""
    
    @bp.route('/projects', methods=['GET'])
    @api_permission_required()
    def get_projects():
        """获取项目列表（分页）"""
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search = request.args.get('search', '')
        status = request.args.get('status', '')
        sort_field = request.args.get('sort_field', 'created_at')
        sort_order = request.args.get('sort_order', 'desc')
        
        current_user_id = getattr(g, 'current_user_id', None)
        # 获取用户信息判断是否为管理员（这里简化处理，实际应从token或数据库获取）
        is_admin = False
        
        result = project_management_repo.get_all(
            page=page,
            per_page=per_page,
            search=search if search else None,
            status=status if status else None,
            sort_field=sort_field,
            sort_order=sort_order,
            current_user_id=current_user_id,
            is_admin=is_admin
        )
        
        return jsonify(result), 200
    
    @bp.route('/projects/<project_id>', methods=['GET'])
    @api_permission_required()
    def get_project(project_id):
        """获取单个项目详情"""
        project = project_management_repo.get_by_id(project_id)
        if not project:
            return jsonify({'error': '项目不存在'}), 404
        return jsonify(project), 200
    
    @bp.route('/projects', methods=['POST'])
    @api_permission_required()
    def create_project():
        """创建项目"""
        data = request.get_json()
        
        required_fields = ['project_no', 'company_name', 'contact_person', 'contact_phone']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} 不能为空'}), 400
        
        current_user_id = getattr(g, 'current_user_id', None)
        
        project = project_management_repo.create(data, current_user_id)
        if not project:
            return jsonify({'error': '项目编号已存在'}), 409
        
        return jsonify(project), 201
    
    @bp.route('/projects/<project_id>', methods=['PUT'])
    @api_permission_required()
    def update_project(project_id):
        """更新项目"""
        data = request.get_json()
        current_user_id = getattr(g, 'current_user_id', None)
        
        project = project_management_repo.update(project_id, data, current_user_id)
        if not project:
            return jsonify({'error': '项目不存在或项目编号已存在'}), 404
        
        return jsonify(project), 200
    
    @bp.route('/projects/<project_id>', methods=['DELETE'])
    @api_permission_required()
    def delete_project(project_id):
        """删除项目"""
        if not project_management_repo.delete(project_id):
            return jsonify({'error': '项目不存在'}), 404
        return '', 204

    @bp.route('/projects/<project_id>/export', methods=['GET'])
    @api_permission_required()
    def export_project_report(project_id):
        """导出项目快测报告为Excel文件（模拟数据版本）"""
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        
        # 删除默认的Sheet
        wb.remove(wb.active)
        
        # 设置样式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ========== 1. 项目信息Sheet ==========
        ws_info = wb.create_sheet("项目信息", 0)
        
        # 表头
        info_headers = ['字段', '值']
        for col, header in enumerate(info_headers, 1):
            cell = ws_info.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 模拟项目数据
        info_data = [
            ['项目编号', 'PJ-2024-001'],
            ['单位名称', 'XX科技有限公司'],
            ['联系人', '张三'],
            ['联系方式', '138****1234'],
            ['项目状态', '进行中'],
            ['资产数量', '15'],
            ['创建时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['备注', '示例数据，请根据实际数据替换'],
        ]
        
        for row_idx, row_data in enumerate(info_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx == 1:
                    cell.font = Font(bold=True)
        
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 40
        
        # ========== 2. 资产清单Sheet ==========
        ws_assets = wb.create_sheet("资产清单")
        
        # 表头
        asset_headers = ['序号', '设备名称', '设备类型', '主机地址', '硬件型号', '软件版本', '虚拟化设备', '域名', '重要程度', '数量', '测评类型']
        for col, header in enumerate(asset_headers, 1):
            cell = ws_assets.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 模拟资产数据
        mock_assets = [
            [1, '核心交换机', '网络设备', '10.0.0.1', 'S12708', 'V200R019', '否', 'switch.example.com', '高', 2, '网络安全测评'],
            [2, '防火墙', '安全设备', '10.0.0.2', 'USG6650', 'V500R005', '否', 'fw.example.com', '高', 1, '边界安全测评'],
            [3, 'Web服务器', '服务器', '10.0.1.10', '华为2288H', 'CentOS 7.9', '是', 'web.example.com', '高', 3, '应用安全测评'],
            [4, '数据库服务器', '服务器', '10.0.1.20', '浪潮NF5280M5', 'MySQL 8.0', '否', 'db.example.com', '高', 1, '数据安全测评'],
            [5, '日志审计系统', '安全设备', '10.0.0.5', 'LogAudit', 'V3.0', '否', 'log.example.com', '中', 1, '日志审计测评'],
        ]
        
        for row_idx, asset in enumerate(mock_assets, 2):
            for col_idx, value in enumerate(asset, 1):
                cell = ws_assets.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 调整列宽
        column_widths = [8, 25, 15, 18, 15, 15, 12, 22, 10, 8, 20]
        for i, width in enumerate(column_widths, 1):
            ws_assets.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # ========== 3. 测评记录Sheet ==========
        ws_records = wb.create_sheet("测评记录")
        
        # 模拟测评指标
        indicators = ['密码长度', '登录失败次数', '锁定时间', '超时时间', '日志功能']
        record_headers = ['设备名称', '设备类型'] + indicators
        
        for col, header in enumerate(record_headers, 1):
            cell = ws_records.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 模拟测评记录数据
        mock_records = [
            ['核心交换机', '网络设备', '8位以上', '5次锁定', '30分钟', '15分钟', '已开启'],
            ['防火墙', '安全设备', '12位以上', '3次锁定', '15分钟', '10分钟', '已开启'],
            ['Web服务器', '服务器', '8位', '5次锁定', '30分钟', '20分钟', '未配置'],
            ['数据库服务器', '服务器', '10位', '3次锁定', '20分钟', '15分钟', '已开启'],
        ]
        
        for row_idx, record in enumerate(mock_records, 2):
            for col_idx, value in enumerate(record, 1):
                cell = ws_records.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        ws_records.column_dimensions['A'].width = 20
        ws_records.column_dimensions['B'].width = 15
        for i in range(3, len(record_headers) + 1):
            ws_records.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"快测报告_{timestamp}.xlsx"
        
        # 返回文件
        response = make_response(send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ))
        
        return response
